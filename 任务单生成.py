import sys
import os
import json
import datetime
from PyQt5.QtWidgets import (QApplication, QInputDialog, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QGridLayout, QLabel, QLineEdit, 
                             QTextEdit, QComboBox, QDateEdit, QSpinBox, 
                             QDoubleSpinBox, QCheckBox, QPushButton, 
                             QTableWidget, QTableWidgetItem, QTabWidget,
                             QGroupBox, QFrame, QSplitter, QMessageBox,
                             QFileDialog, QProgressBar, QToolBar, QAction,
                             QStatusBar, QMenu, QMenuBar, QListWidget,
                             QTreeWidget, QTreeWidgetItem, QHeaderView,
                             QDialog, QDialogButtonBox, QFormLayout)
from PyQt5.QtCore import Qt, QDate, QTimer, pyqtSignal, QThread, QSettings
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor, QPixmap, QPainter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
import sqlite3
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DatabaseManager:
    """数据库管理类"""
    
    def __init__(self, db_path="tasks.db"):
        self.db_path = db_path
        self.init_db()
    
    def init_db(self):
        """初始化数据库"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 创建任务表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT,
                priority INTEGER DEFAULT 1,
                status TEXT DEFAULT 'pending',
                assignee TEXT,
                created_date TEXT,
                due_date TEXT,
                completed_date TEXT,
                category TEXT,
                estimated_hours REAL,
                actual_hours REAL,
                tags TEXT
            )
        ''')
        
        # 创建用户表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                full_name TEXT,
                email TEXT,
                role TEXT DEFAULT 'user'
            )
        ''')
        
        # 创建类别表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                color TEXT DEFAULT '#FFFFFF'
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def execute_query(self, query, params=None, fetch=False):
        """执行查询"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            if fetch:
                result = cursor.fetchall()
                conn.close()
                return result
            else:
                conn.commit()
                conn.close()
                return cursor
                
        except Exception as e:
            logger.error(f"数据库查询错误: {e}")
            conn.close()
            return None
    
    def fetch_all(self, query, params=None):
        """获取所有结果"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            result = cursor.fetchall()
            return result
        except Exception as e:
            logger.error(f"数据库查询错误: {e}")
            return []
        finally:
            conn.close()
    
    def fetch_one(self, query, params=None):
        """获取单个结果"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchone()
        except Exception as e:
            logger.error(f"数据库查询错误: {e}")
            return None
        finally:
            conn.close()

class TaskManager:
    """任务管理类"""
    
    def __init__(self, db_manager):
        self.db = db_manager
    
    def add_task(self, task_data):
        """添加新任务"""
        query = '''
            INSERT INTO tasks (title, description, priority, status, assignee, 
                              created_date, due_date, category, estimated_hours, tags)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        '''
        params = (
            task_data.get('title'),
            task_data.get('description'),
            task_data.get('priority', 1),
            task_data.get('status', 'pending'),
            task_data.get('assignee'),
            datetime.datetime.now().isoformat(),
            task_data.get('due_date'),
            task_data.get('category'),
            task_data.get('estimated_hours', 0),
            json.dumps(task_data.get('tags', []))
        )
        return self.db.execute_query(query, params)
    
    def update_task(self, task_id, task_data):
        """更新任务"""
        query = '''
            UPDATE tasks 
            SET title=?, description=?, priority=?, status=?, assignee=?, 
                due_date=?, category=?, estimated_hours=?, actual_hours=?, tags=?
            WHERE id=?
        '''
        params = (
            task_data.get('title'),
            task_data.get('description'),
            task_data.get('priority'),
            task_data.get('status'),
            task_data.get('assignee'),
            task_data.get('due_date'),
            task_data.get('category'),
            task_data.get('estimated_hours'),
            task_data.get('actual_hours'),
            json.dumps(task_data.get('tags', [])),
            task_id
        )
        return self.db.execute_query(query, params)
    
    def get_task(self, task_id):
        """获取任务详情"""
        query = "SELECT * FROM tasks WHERE id=?"
        result = self.db.fetch_one(query, (task_id,))
        if result:
            return self._format_task(result)
        return None
    
    def get_all_tasks(self, filters=None):
        """获取所有任务（可过滤）"""
        query = "SELECT * FROM tasks WHERE 1=1"
        params = []
        
        if filters:
            if filters.get('status'):
                query += " AND status=?"
                params.append(filters['status'])
            if filters.get('priority'):
                query += " AND priority=?"
                params.append(filters['priority'])
            if filters.get('assignee'):
                query += " AND assignee=?"
                params.append(filters['assignee'])
            if filters.get('category'):
                query += " AND category=?"
                params.append(filters['category'])
        
        query += " ORDER BY priority DESC, due_date ASC"
        
        results = self.db.fetch_all(query, params)
        return [self._format_task(row) for row in results]
    
    def _format_task(self, row):
        """格式化任务数据"""
        return {
            'id': row[0],
            'title': row[1],
            'description': row[2],
            'priority': row[3],
            'status': row[4],
            'assignee': row[5],
            'created_date': row[6],
            'due_date': row[7],
            'completed_date': row[8],
            'category': row[9],
            'estimated_hours': row[10],
            'actual_hours': row[11],
            'tags': json.loads(row[12]) if row[12] else []
        }

class ExportManager:
    """导出管理类"""
    
    @staticmethod
    def export_to_pdf(tasks, filename):
        """导出任务到PDF"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # 标题
            title = Paragraph("任务单报告", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # 生成日期
            date_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            date_para = Paragraph(f"生成日期: {date_str}", styles['Normal'])
            elements.append(date_para)
            elements.append(Spacer(1, 12))
            
            # 任务表格
            if tasks:
                data = [['ID', '标题', '优先级', '状态', '负责人', '截止日期']]
                for task in tasks:
                    data.append([
                        str(task['id']),
                        task['title'],
                        str(task['priority']),
                        task['status'],
                        task['assignee'] or '',
                        task['due_date'] or ''
                    ])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph("没有任务数据", styles['Normal']))
            
            doc.build(elements)
            return True
        except Exception as e:
            logger.error(f"PDF导出错误: {e}")
            return False
    
    @staticmethod
    def export_to_excel(tasks, filename):
        """导出任务到Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"
            
            # 标题行
            headers = ['ID', '标题', '描述', '优先级', '状态', '负责人', '创建日期', '截止日期', '类别']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 数据行
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task['id'])
                ws.cell(row=row, column=2, value=task['title'])
                ws.cell(row=row, column=3, value=task['description'])
                ws.cell(row=row, column=4, value=task['priority'])
                ws.cell(row=row, column=5, value=task['status'])
                ws.cell(row=row, column=6, value=task['assignee'])
                ws.cell(row=row, column=7, value=task['created_date'])
                ws.cell(row=row, column=8, value=task['due_date'])
                ws.cell(row=row, column=9, value=task['category'])
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return True
        except Exception as e:
            logger.error(f"Excel导出错误: {e}")
            return False

class EmailManager:
    """邮件管理类"""
    
    def __init__(self, smtp_server, port, username, password):
        self.smtp_server = smtp_server
        self.port = port
        self.username = username
        self.password = password
    
    def send_task_report(self, recipient, tasks, subject="任务单报告"):
        """发送任务报告邮件"""
        try:
            # 创建邮件
            msg = MIMEMultipart()
            msg['From'] = self.username
            msg['To'] = recipient
            msg['Subject'] = subject
            
            # 邮件正文
            body = f"""
            <h2>任务单报告</h2>
            <p>生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
            <p>任务总数: {len(tasks)}</p>
            <table border="1" style="border-collapse: collapse;">
                <tr>
                    <th>ID</th><th>标题</th><th>优先级</th><th>状态</th><th>负责人</th>
                </tr>
            """
            
            for task in tasks:
                body += f"""
                <tr>
                    <td>{task['id']}</td>
                    <td>{task['title']}</td>
                    <td>{task['priority']}</td>
                    <td>{task['status']}</td>
                    <td>{task['assignee'] or '未分配'}</td>
                </tr>
                """
            
            body += "</table>"
            
            msg.attach(MIMEText(body, 'html'))
            
            # 连接服务器并发送
            server = smtplib.SMTP(self.smtp_server, self.port)
            server.starttls()
            server.login(self.username, self.password)
            server.send_message(msg)
            server.quit()
            
            return True
        except Exception as e:
            logger.error(f"邮件发送错误: {e}")
            return False

class CustomWidgets:
    """自定义控件类"""
    
    @staticmethod
    def create_priority_combo():
        """创建优先级下拉框"""
        combo = QComboBox()
        combo.addItem("低", 1)
        combo.addItem("中", 2)
        combo.addItem("高", 3)
        combo.addItem("紧急", 4)
        return combo
    
    @staticmethod
    def create_status_combo():
        """创建状态下拉框"""
        combo = QComboBox()
        combo.addItem("待处理", "pending")
        combo.addItem("进行中", "in_progress")
        combo.addItem("已完成", "completed")
        combo.addItem("已取消", "cancelled")
        return combo
    
    @staticmethod
    def create_tag_editor():
        """创建标签编辑器"""
        class TagEditor(QWidget):
            tagChanged = pyqtSignal(list)
            
            def __init__(self):
                super().__init__()
                self.tags = []
                self.init_ui()
            
            def init_ui(self):
                layout = QHBoxLayout()
                self.setLayout(layout)
                
                self.tag_input = QLineEdit()
                self.tag_input.setPlaceholderText("输入标签并按回车添加")
                self.tag_input.returnPressed.connect(self.add_tag)
                layout.addWidget(self.tag_input)
                
                self.add_btn = QPushButton("添加")
                self.add_btn.clicked.connect(self.add_tag)
                layout.addWidget(self.add_btn)
                
                self.tags_layout = QHBoxLayout()
                layout.addLayout(self.tags_layout)
            
            def add_tag(self):
                tag_text = self.tag_input.text().strip()
                if tag_text and tag_text not in self.tags:
                    self.tags.append(tag_text)
                    self.update_tags_display()
                    self.tag_input.clear()
            
            def remove_tag(self, tag):
                if tag in self.tags:
                    self.tags.remove(tag)
                    self.update_tags_display()
            
            def update_tags_display(self):
                # 清除现有标签显示
                for i in reversed(range(self.tags_layout.count())):
                    widget = self.tags_layout.itemAt(i).widget()
                    if widget:
                        widget.deleteLater()
                
                # 添加标签按钮
                for tag in self.tags:
                    tag_btn = QPushButton(f"× {tag}")
                    tag_btn.setStyleSheet("QPushButton { border: none; color: blue; }")
                    tag_btn.clicked.connect(lambda checked, t=tag: self.remove_tag(t))
                    self.tags_layout.addWidget(tag_btn)
                
                self.tagChanged.emit(self.tags)
            
            def set_tags(self, tags):
                self.tags = tags
                self.update_tags_display()
            
            def get_tags(self):
                return self.tags
        
        return TagEditor()

class TaskDialog(QDialog):
    """任务编辑对话框"""
    
    def __init__(self, task_manager, task_id=None, parent=None):
        super().__init__(parent)
        self.task_manager = task_manager
        self.task_id = task_id
        self.task_data = {}
        
        self.init_ui()
        self.load_task_data()
    
    def init_ui(self):
        self.setWindowTitle("编辑任务" if self.task_id else "新建任务")
        self.setModal(True)
        self.resize(500, 400)
        
        layout = QFormLayout()
        self.setLayout(layout)
        
        # 标题
        self.title_edit = QLineEdit()
        layout.addRow("标题:", self.title_edit)
        
        # 描述
        self.desc_edit = QTextEdit()
        self.desc_edit.setMaximumHeight(100)
        layout.addRow("描述:", self.desc_edit)
        
        # 优先级
        self.priority_combo = CustomWidgets.create_priority_combo()
        layout.addRow("优先级:", self.priority_combo)
        
        # 状态
        self.status_combo = CustomWidgets.create_status_combo()
        layout.addRow("状态:", self.status_combo)
        
        # 负责人
        self.assignee_edit = QLineEdit()
        layout.addRow("负责人:", self.assignee_edit)
        
        # 截止日期
        self.due_date_edit = QDateEdit()
        self.due_date_edit.setDate(QDate.currentDate().addDays(7))
        self.due_date_edit.setCalendarPopup(True)
        layout.addRow("截止日期:", self.due_date_edit)
        
        # 类别
        self.category_edit = QLineEdit()
        layout.addRow("类别:", self.category_edit)
        
        # 预估工时
        self.estimated_hours_spin = QDoubleSpinBox()
        self.estimated_hours_spin.setRange(0, 1000)
        self.estimated_hours_spin.setSuffix(" 小时")
        layout.addRow("预估工时:", self.estimated_hours_spin)
        
        # 标签
        self.tag_editor = CustomWidgets.create_tag_editor()
        layout.addRow("标签:", self.tag_editor)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addRow(button_box)
    
    def load_task_data(self):
        if self.task_id:
            self.task_data = self.task_manager.get_task(self.task_id)
            if self.task_data:
                self.title_edit.setText(self.task_data.get('title', ''))
                self.desc_edit.setPlainText(self.task_data.get('description', ''))
                
                # 设置优先级
                index = self.priority_combo.findData(self.task_data.get('priority', 1))
                if index >= 0:
                    self.priority_combo.setCurrentIndex(index)
                
                # 设置状态
                index = self.status_combo.findData(self.task_data.get('status', 'pending'))
                if index >= 0:
                    self.status_combo.setCurrentIndex(index)
                
                self.assignee_edit.setText(self.task_data.get('assignee', ''))
                
                # 设置截止日期
                if self.task_data.get('due_date'):
                    due_date = QDate.fromString(self.task_data['due_date'][:10], 'yyyy-MM-dd')
                    self.due_date_edit.setDate(due_date)
                
                self.category_edit.setText(self.task_data.get('category', ''))
                self.estimated_hours_spin.setValue(self.task_data.get('estimated_hours', 0))
                self.tag_editor.set_tags(self.task_data.get('tags', []))
    
    def get_task_data(self):
        return {
            'title': self.title_edit.text(),
            'description': self.desc_edit.toPlainText(),
            'priority': self.priority_combo.currentData(),
            'status': self.status_combo.currentData(),
            'assignee': self.assignee_edit.text(),
            'due_date': self.due_date_edit.date().toString('yyyy-MM-dd'),
            'category': self.category_edit.text(),
            'estimated_hours': self.estimated_hours_spin.value(),
            'tags': self.tag_editor.get_tags()
        }

class MainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        
        # 初始化管理器
        self.db_manager = DatabaseManager()
        self.task_manager = TaskManager(self.db_manager)
        self.export_manager = ExportManager()
        
        # 从设置加载邮件配置
        self.settings = QSettings("MyCompany", "TaskManager")
        self.email_manager = None
        self.init_email_manager()
        
        self.init_ui()
        self.load_tasks()
    
    def init_email_manager(self):
        """初始化邮件管理器"""
        smtp_server = self.settings.value("smtp_server", "")
        port = self.settings.value("smtp_port", 587, type=int)
        username = self.settings.value("email_username", "")
        password = self.settings.value("email_password", "")
        
        if smtp_server and username and password:
            self.email_manager = EmailManager(smtp_server, port, username, password)
    
    def init_ui(self):
        self.setWindowTitle("高级任务单生成系统")
        self.setGeometry(100, 100, 1200, 800)
        
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 创建工具栏
        self.create_toolbar()
        
        # 创建过滤栏
        self.create_filter_bar()
        
        # 创建任务表格
        self.create_task_table()
        
        # 创建状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")
    
    def create_toolbar(self):
        toolbar = QToolBar()
        self.addToolBar(toolbar)
        
        # 新建任务动作
        new_action = QAction("新建任务", self)
        new_action.triggered.connect(self.new_task)
        toolbar.addAction(new_action)
        
        # 编辑任务动作
        edit_action = QAction("编辑任务", self)
        edit_action.triggered.connect(self.edit_task)
        toolbar.addAction(edit_action)
        
        toolbar.addSeparator()
        
        # 导出PDF动作
        export_pdf_action = QAction("导出PDF", self)
        export_pdf_action.triggered.connect(self.export_pdf)
        toolbar.addAction(export_pdf_action)
        
        # 导出Excel动作
        export_excel_action = QAction("导出Excel", self)
        export_excel_action.triggered.connect(self.export_excel)
        toolbar.addAction(export_excel_action)
        
        toolbar.addSeparator()
        
        # 发送邮件动作
        email_action = QAction("发送报告", self)
        email_action.triggered.connect(self.send_email_report)
        toolbar.addAction(email_action)
        
        # 设置动作
        settings_action = QAction("设置", self)
        settings_action.triggered.connect(self.show_settings)
        toolbar.addAction(settings_action)
    
    def create_filter_bar(self):
        filter_widget = QWidget()
        filter_layout = QHBoxLayout()
        filter_widget.setLayout(filter_layout)
        
        # 状态过滤器
        filter_layout.addWidget(QLabel("状态:"))
        self.status_filter = CustomWidgets.create_status_combo()
        self.status_filter.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.status_filter)
        
        # 优先级过滤器
        filter_layout.addWidget(QLabel("优先级:"))
        self.priority_filter = CustomWidgets.create_priority_combo()
        self.priority_filter.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.priority_filter)
        
        # 负责人过滤器
        filter_layout.addWidget(QLabel("负责人:"))
        self.assignee_filter = QLineEdit()
        self.assignee_filter.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.assignee_filter)
        
        # 搜索框
        filter_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索任务标题或描述...")
        self.search_edit.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.search_edit)
        
        filter_layout.addStretch()
        
        self.layout().addWidget(filter_widget)
    
    def create_task_table(self):
        self.task_table = QTableWidget()
        self.task_table.setColumnCount(8)
        self.task_table.setHorizontalHeaderLabels([
            "ID", "标题", "优先级", "状态", "负责人", "截止日期", "类别", "标签"
        ])
        
        # 设置表格属性
        self.task_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.task_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.task_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.task_table.doubleClicked.connect(self.edit_task)
        
        self.layout().addWidget(self.task_table)
    
    def load_tasks(self, filters=None):
        """加载任务到表格"""
        tasks = self.task_manager.get_all_tasks(filters)
        
        self.task_table.setRowCount(len(tasks))
        
        for row, task in enumerate(tasks):
            self.task_table.setItem(row, 0, QTableWidgetItem(str(task['id'])))
            self.task_table.setItem(row, 1, QTableWidgetItem(task['title']))
            self.task_table.setItem(row, 2, QTableWidgetItem(str(task['priority'])))
            self.task_table.setItem(row, 3, QTableWidgetItem(task['status']))
            self.task_table.setItem(row, 4, QTableWidgetItem(task['assignee'] or ''))
            self.task_table.setItem(row, 5, QTableWidgetItem(task['due_date'] or ''))
            self.task_table.setItem(row, 6, QTableWidgetItem(task['category'] or ''))
            self.task_table.setItem(row, 7, QTableWidgetItem(', '.join(task['tags'])))
        
        self.status_bar.showMessage(f"加载了 {len(tasks)} 个任务")
    
    def apply_filters(self):
        """应用过滤器"""
        filters = {}
        
        status = self.status_filter.currentData()
        if status != "pending":  # 默认显示所有
            filters['status'] = status
        
        priority = self.priority_filter.currentData()
        if priority != 1:  # 默认显示所有
            filters['priority'] = priority
        
        assignee = self.assignee_filter.text().strip()
        if assignee:
            filters['assignee'] = assignee
        
        self.load_tasks(filters)
    
    def new_task(self):
        """新建任务"""
        dialog = TaskDialog(self.task_manager)
        if dialog.exec_() == QDialog.Accepted:
            task_data = dialog.get_task_data()
            self.task_manager.add_task(task_data)
            self.load_tasks()
            self.status_bar.showMessage("任务创建成功")
    
    def edit_task(self):
        """编辑任务"""
        selected_row = self.task_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一个任务")
            return
        
        task_id = int(self.task_table.item(selected_row, 0).text())
        dialog = TaskDialog(self.task_manager, task_id, self)
        if dialog.exec_() == QDialog.Accepted:
            task_data = dialog.get_task_data()
            self.task_manager.update_task(task_id, task_data)
            self.load_tasks()
            self.status_bar.showMessage("任务更新成功")
    
    def export_pdf(self):
        """导出为PDF"""
        filename, _ = QFileDialog.getSaveFileName(
            self, "导出PDF", "tasks.pdf", "PDF文件 (*.pdf)")
        
        if filename:
            tasks = self.task_manager.get_all_tasks()
            if self.export_manager.export_to_pdf(tasks, filename):
                QMessageBox.information(self, "成功", f"PDF已导出到: {filename}")
            else:
                QMessageBox.warning(self, "错误", "PDF导出失败")
    
    def export_excel(self):
        """导出为Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", "tasks.xlsx", "Excel文件 (*.xlsx)")
        
        if filename:
            tasks = self.task_manager.get_all_tasks()
            if self.export_manager.export_to_excel(tasks, filename):
                QMessageBox.information(self, "成功", f"Excel已导出到: {filename}")
            else:
                QMessageBox.warning(self, "错误", "Excel导出失败")
    
    def send_email_report(self):
        """发送邮件报告"""
        if not self.email_manager:
            QMessageBox.warning(self, "错误", "请先配置邮件设置")
            self.show_settings()
            return
        
        recipient, ok = QInputDialog.getText(self, "发送报告", "请输入收件人邮箱:")
        if ok and recipient:
            tasks = self.task_manager.get_all_tasks()
            if self.email_manager.send_task_report(recipient, tasks):
                QMessageBox.information(self, "成功", "邮件发送成功")
            else:
                QMessageBox.warning(self, "错误", "邮件发送失败")
    
    def show_settings(self):
        """显示设置对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("设置")
        dialog.resize(400, 300)
        
        layout = QFormLayout()
        dialog.setLayout(layout)
        
        # 邮件设置
        layout.addWidget(QLabel("<h3>邮件设置</h3>"))
        
        smtp_server_edit = QLineEdit(self.settings.value("smtp_server", ""))
        layout.addRow("SMTP服务器:", smtp_server_edit)
        
        smtp_port_spin = QSpinBox()
        smtp_port_spin.setRange(1, 65535)
        smtp_port_spin.setValue(self.settings.value("smtp_port", 587, type=int))
        layout.addRow("SMTP端口:", smtp_port_spin)
        
        email_username_edit = QLineEdit(self.settings.value("email_username", ""))
        layout.addRow("邮箱用户名:", email_username_edit)
        
        email_password_edit = QLineEdit(self.settings.value("email_password", ""))
        email_password_edit.setEchoMode(QLineEdit.Password)
        layout.addRow("邮箱密码:", email_password_edit)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)
        
        if dialog.exec_() == QDialog.Accepted:
            # 保存设置
            self.settings.setValue("smtp_server", smtp_server_edit.text())
            self.settings.setValue("smtp_port", smtp_port_spin.value())
            self.settings.setValue("email_username", email_username_edit.text())
            self.settings.setValue("email_password", email_password_edit.text())
            
            # 重新初始化邮件管理器
            self.init_email_manager()
            
            QMessageBox.information(self, "成功", "设置已保存")

# 应用程序入口
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # 设置应用程序样式
    app.setStyle("Fusion")
    
    # 创建并显示主窗口
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())